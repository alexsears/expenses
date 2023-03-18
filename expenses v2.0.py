import pandas as pd
import tkinter as tk
from tkinter import filedialog
import matplotlib.pyplot as plt
import calendar
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter




def create_cover_sheet(tests, plots, descriptions):
    # Create a new workbook
    #wb = openpyxl.Workbook()

    # Create the cover sheet
    sheet = wb.active
    sheet.title = "Cover Sheet"

    # Add a title
    sheet['A1'] = "List of Tests and Plots"

    # Add the test names and plot links
    current_row = 3
    for i, test in enumerate(tests):
        # Add the test name as a hyperlink to the corresponding sheet, must be formatted blue and underlined
        sheet[f'A{current_row}'].value = f"Test {i+1}: {test}"
        sheet[f'A{current_row}'].hyperlink = f"'Test {i+1}'!A1"
        current_row += 1

    # Add the plot links
    for i, plot in enumerate(plots):
        # Add the plot name as a hyperlink to the corresponding sheet
        sheet[f'A{current_row}'].value = f"Plot {i+1}: {plot}"
        sheet[f'A{current_row}'].hyperlink = f"'Plot {i+1}'!A1"
        current_row += 1

    # Resize columns to fit data
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    # Add descriptions to all tabs and add links back to the cover sheet
    for ws in wb.worksheets:
        if ws.title.startswith("Test "):
            test_num = int(ws.title.split()[1])
            ws['A1'] = f"Test {test_num}: {tests[test_num - 1]}"
        elif ws.title.startswith("Plot "):
            plot_num = int(ws.title.split()[1])
            ws['A1'] = f"Plot {plot_num}: {plots[plot_num - 1]}"

            # Add a link back to the cover sheet
            
        ws['A1'].hyperlink = "'Cover Sheet'!A1"

    return wb

def within_1_percent(df):
    """
    Filter data within 1% of each other in the Approved Amount 
    """
    if len(df) >= 2:
        mean_payment = df['Approved Amount (rpt)'].mean()
        upper_limit = mean_payment * 1.01
        lower_limit = mean_payment * 0.99
        return df[(df['Approved Amount (rpt)'] <= upper_limit) & (df['Approved Amount (rpt)'] >= lower_limit)]
    else:
        return pd.DataFrame()
def within_5_days(df):
    """
    Filter data for transactions that happened within 5 days of each other
    """
    if len(df) >= 2:
        max_date = df['Transaction Date'].max()
        min_date = df['Transaction Date'].min()
        if (max_date - min_date).days <= 5:
            return df
        else:
            return pd.DataFrame()
    else:
        return pd.DataFrame()
# Define a function to filter groups with less than 2 rows
def at_least_2_rows(df):
    return len(df) >= 2
def select_file():
    root = tk.Tk()
    root.withdraw()

    while True:
        print("Select a file to load for analysis")
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xls;*.xlsx;*.xlsm;*.xlsb;*.odf;*.ods;*.odt")])

        if file_path:
            break
        else:
            print("No file selected. Please try again.")

    return file_path

file_path = select_file()
df = pd.read_excel(file_path, skiprows=2, engine='openpyxl')

def select_save_file():
    """
    Opens a Windows dialog for the user to select a file name and location
    to save the results as an Excel file.

    Returns:
        str: The selected file path, or None if no file was selected.
    """
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        title="Select a file to save the results",
        filetypes=[("Excel Files", "*.xlsx")]
    )
    return file_path if file_path else None

# Function to save DataFrame to Excel sheet
def save_df_to_sheet(workbook, sheet_name, df):
    ws = workbook.create_sheet(sheet_name)
    for r in dataframe_to_rows(df, index=True, header=True):
        ws.append(r)
    
    # Resize columns to fit data
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width


# Function to save plot to Excel sheet
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image

def save_plot_to_sheet(workbook, sheet_name, plot_data, img_name):
    fig, ax = plt.subplots()
    # plot data, title = test or plot description
    plot_data.plot(kind='bar', title='', ax=ax)
    fig.savefig(img_name, dpi=100)
    ws = workbook.create_sheet(sheet_name)
    img = Image(img_name)
    ws.add_image(img, 'A2')
    # save the dataframe starting in cell L1
    rows = dataframe_to_rows(plot_data.reset_index(), index=False, header=True)

    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 17):  # Change 1 to 12 to start from the L column
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = value
    # Resize columns to fit data
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width


        
    

# Example usage:
selected_file = select_save_file()
# Create a new workbook
wb = Workbook()
if selected_file:
    print(f"Selected file: {selected_file}")
else:
    print("No file selected.")

# Convert the columns to formats we need
df['Transaction Date'] = pd.to_datetime(df['Transaction Date'])
df['Claimed Amount'] = pd.to_numeric(df['Claimed Amount'])

# approved amount is a string, so we need to convert it to a currency with 2 decimal places and commas
df['Approved Amount (rpt)'] = pd.to_numeric(df['Approved Amount (rpt)'])



df['Sent for Payment Date'] = pd.to_datetime(df['Sent for Payment Date'])

# Filter data for transactions over 1000 and sort by employee ID  
filtered_df = df[df['Approved Amount (rpt)'] > 1].sort_values(by=['Employee ID'])

#sort by employee ID
filtered_df = filtered_df.sort_values(by=['Employee ID'])

#limit to 100. rows
#filtered_df = filtered_df.head(1000)

# TEST 1-------------------Same employee, same vendor same total
#define test1
test1 = 'Same employee, same vendor same total'
# Group by employee id, vendor, and approved amount
grouped_df = filtered_df.groupby(['Employee ID', 'Vendor', 'Approved Amount (rpt)'])

# Filter groups with more than one row per employee
filtered_groups = grouped_df.filter(lambda x: len(x) > 1)

# Sort the filtered groups by employee id, vendor, and approved amount
sorted_filtered_groups = filtered_groups.sort_values(by=['Employee ID', 'Vendor', 'Approved Amount (rpt)', 'Transaction Date'])

# Display tuples with each other, separated by a blank line
#with pd.option_context('display.max_rows', None, 'display.max_columns', None):
#    for _, group in sorted_filtered_groups.groupby(['Employee ID', 'Vendor', 'Approved Amount (rpt)']):
#        print(group[['Employee ID', 'Vendor', 'Approved Amount (rpt)', 'Transaction Date']])
#        print("")

# Print the total number of rows
#print("Total number of rows:", len(sorted_filtered_groups))

# Print the first 50 rows of the grouped data same vendor,
#print("Test 1 - same employee, vendor, and total")
#print(sorted_filtered_groups[['Employee ID', 'Vendor', 'Approved Amount (rpt)', 'Transaction Date']].head(50))

save_df_to_sheet(wb, "Test 1", sorted_filtered_groups[['Employee ID', 'Vendor', 'Approved Amount (rpt)', 'Transaction Date']])

# TEST 2-----------same employee, amount and date -----------------------------

#define test2
test2 = 'Same employee, amount and date'

# Group by employee id, approved amount, and transaction date
grouped_df2 = filtered_df.groupby(['Employee ID', 'Approved Amount (rpt)', 'Transaction Date'])

# Filter groups with more than one row per employee
filtered_groups2 = grouped_df2.filter(lambda x: len(x) > 1)

# Sort the filtered groups by employee id, approved amount, and transaction date
sorted_filtered_groups2 = filtered_groups2.sort_values(by=['Employee ID', 'Approved Amount (rpt)', 'Transaction Date'])

# Display tuples with each other, separated by a blank line
#with pd.option_context('display.max_rows', None, 'display.max_columns', None):
#    for _, group in sorted_filtered_groups2.groupby(['Employee ID', 'Approved Amount (rpt)', 'Transaction Date']):
#        print(group[['Employee ID', 'Approved Amount (rpt)', 'Transaction Date', 'Report Name']])
#        print("")

# Print the total number of rows
#print("Total number of rows:", len(sorted_filtered_groups2))

# Print the first 50 rows of the grouped data same amount and date
#print("Test 2 - same employee, amount, and date")
#print(sorted_filtered_groups2[['Employee ID', 'Approved Amount (rpt)', 'Transaction Date', 'Report Name']].head(50))

save_df_to_sheet(wb, "Test 2", sorted_filtered_groups2[['Employee ID', 'Approved Amount (rpt)', 'Transaction Date', 'Report Name']])

# TEST 3-----------same employee, SAME AMOUNT-----------------------------

#define test3
test3 = 'Same employee, same amount'

# Group by employee id and approved amount
grouped_df3 = filtered_df.groupby(['Employee ID', 'Approved Amount (rpt)'])

# Filter groups with more than one row per employee
filtered_groups3 = grouped_df3.filter(lambda x: len(x) > 1)

# Sort the filtered groups by employee id and approved amount
sorted_filtered_groups3 = filtered_groups3.sort_values(by=['Employee ID', 'Approved Amount (rpt)'])

# Display tuples with each other, separated by a blank line
#with pd.option_context('display.max_rows', None, 'display.max_columns', None):
#    for _, group in sorted_filtered_groups3.groupby(['Employee ID', 'Approved Amount (rpt)']):
#        print(group[['Employee ID', 'Approved Amount (rpt)', 'Transaction Date', 'Vendor']])
#        print("")

# Print the total number of rows
#print("Total number of rows:", len(sorted_filtered_groups3))

# Print the first 50 rows of the grouped data same amount
#print("Test 3 - same employee and amount")
#print(sorted_filtered_groups3[['Employee ID', 'Approved Amount (rpt)', 'Transaction Date', 'Vendor']].head(50))

save_df_to_sheet(wb, "Test 3", sorted_filtered_groups3[['Employee ID', 'Approved Amount (rpt)', 'Transaction Date', 'Vendor']])

# TEST 4-----------keyword test-----------------------------
#define test4
test4 = 'Keyword test'


# Create a Tkinter file dialog box to open an Excel file
root = tk.Tk()
root.withdraw()
print("Please select keywords file")     
file_path = filedialog.askopenfilename()

# Read in keywords from the Excel file
keywords = pd.read_excel(file_path)

# Convert the first column of keywords DataFrame to a list
keywords = keywords.iloc[:, 0].tolist()

# Group the data with any of the keywords, sort by keyword, and add a line break
grouped_df4 = filtered_df[filtered_df['Report Name'].str.contains('|'.join(keywords))].sort_values(by=['Report Name'])
                                                                                                                                         
# Remove empty dataframes from the grouped data
grouped_df4 = grouped_df4[grouped_df4['Approved Amount (rpt)'].notna()]

# Print total number of rows
#print("Total number of rows:", len(grouped_df4))

# Print first 50 rows of the grouped data just employee id, amount, date, and report name
#print("Test 4 - Keyword search")
#print(grouped_df4[['Employee ID', 'Transaction Date', 'Report Name', 'Approved Amount (rpt)']].head(50))

save_df_to_sheet(wb, "Test 4", grouped_df4[['Employee ID', 'Transaction Date', 'Report Name', 'Approved Amount (rpt)']])

# TEST 5: Group by same employee and same vendor

# Define test5
test5 = 'Same employee and same vendor'

grouped_df5 = filtered_df.groupby(['Employee ID', 'Vendor']).apply(lambda x: within_1_percent(within_5_days(x)))

# Reset the index after applying the filter functions
grouped_df5.reset_index(drop=True, inplace=True)

# Filter out groups with less than 2 rows
grouped_df5 = grouped_df5[grouped_df5.groupby(['Employee ID', 'Vendor']).transform(at_least_2_rows)['Approved Amount (rpt)']]

# Reset the index again
grouped_df5.reset_index(drop=True, inplace=True)

# Print the first 50 rows of the grouped data with Employee ID, Vendor Name, Transaction Date, Report Name, Transaction Count, and Total Amount
#print("Test 5 - Group by same employee and same vendor, fuzzy match on date and amount")

# Display tuples separated by a blank line
unique_employee_vendor_pairs = grouped_df5[['Employee ID', 'Vendor']].drop_duplicates()

for index, row in unique_employee_vendor_pairs.iterrows():
    employee_id = row['Employee ID']
    vendor = row['Vendor']
    temp_df = grouped_df5[(grouped_df5['Employee ID'] == employee_id) & (grouped_df5['Vendor'] == vendor)]
    print(temp_df[['Employee ID', 'Vendor', 'Transaction Date', 'Report Name', 'Approved Amount (rpt)']])
    print("\n")

# Save the grouped data to a new sheet in the Excel file
save_df_to_sheet(wb, "Test 5", grouped_df5[['Employee ID', 'Vendor', 'Transaction Date', 'Report Name', 'Approved Amount (rpt)']])

# TEST 6: Group by same employee and expense type
# Define test6
test6 = 'Same employee and expense type'

grouped_df6 = filtered_df.groupby(['Employee ID', 'Expense Type']).apply(lambda x: within_1_percent(within_5_days(x)))

# Reset the index after applying the filter functions
grouped_df6.reset_index(drop=True, inplace=True)

# Filter out groups with less than 2 rows
grouped_df6 = grouped_df6[grouped_df6.groupby(['Employee ID', 'Expense Type']).transform(at_least_2_rows)['Approved Amount (rpt)']]

# Reset the index again
grouped_df6.reset_index(drop=True, inplace=True)

# Print total number of rows
#print("Total number of rows:", len(grouped_df6))

# Print the first 50 rows of the grouped data with Employee ID, Expense Type, Transaction Date, Report Name, Transaction Count, and Total Amount
#print("Test 6 - Group by same employee and expense type")

# Display tuples separated by a blank line
unique_employee_expense_pairs = grouped_df6[['Employee ID', 'Expense Type']].drop_duplicates()

for index, row in unique_employee_expense_pairs.iterrows():
    employee_id = row['Employee ID']
    expense_type = row['Expense Type']
    temp_df = grouped_df6[(grouped_df6['Employee ID'] == employee_id) & (grouped_df6['Expense Type'] == expense_type)]
#    print(temp_df[['Employee ID', 'Expense Type', 'Transaction Date', 'Report Name', 'Approved Amount (rpt)']])
#    print("\n")

# Save the grouped data to a new sheet in the Excel file
save_df_to_sheet(wb, "Test 6", grouped_df6[['Employee ID', 'Expense Type', 'Transaction Date', 'Report Name', 'Approved Amount (rpt)']])




# print title and a blank line
#print("Plot 1 - Top 20 Employees by Total Approved Amount")
#print()

# define plot 1
plot1 = 'Top 20 Employees by Total Approved Amount'


# print the top 20 employees by total approved amount
top_20_employees = filtered_df.groupby(['Employee'])['Approved Amount (rpt)'].sum().sort_values(ascending=False).head(20)
#print(top_20_employees)



#plot the data by employee id sort by total approved amount with title, top 20 employees
filtered_df.groupby(['Employee'])['Approved Amount (rpt)'].sum().sort_values(ascending=False).head(20).plot(kind='bar', title='Top 20 Employees by Total Approved Amount')

# Create plot data
plot_data = filtered_df.groupby(['Employee'])['Approved Amount (rpt)'].sum().sort_values(ascending=False).head(20)

# Save the plot to Excel sheet
save_plot_to_sheet(wb, "Plot 1", plot_data, "temp_plot1.png")

# print title and a blank line
#print("Plot 2 - Top 20 Vendors by Total Approved Amount")
#print()
# define plot 2
plot2 = 'Top 20 Vendors by Total Approved Amount'

# Create the top 20 vendors dataframe
top_20_vendors = filtered_df.groupby(['Vendor'])['Approved Amount (rpt)'].sum().sort_values(ascending=False).head(20)

# Plot the data by vendor and approved amount sort by total approved amount with title top 20 vendors
top_20_vendors.plot(kind='bar', title='Top 20 Vendors by Total Approved Amount')

# Save the plot to Excel sheet
save_plot_to_sheet(wb, "Plot 2", top_20_vendors, "temp_plot2.png")



# print title and a blank line
#print("Plot 3 - Top 20 Report Names by Total Approved Amount")
#print()

# define plot 3
plot3 = 'Top 20 Report Names by Total Approved Amount'

# Create the top 20 report names dataframe
top_20_report_names = filtered_df.groupby(['Employee', 'Report Name'])['Approved Amount (rpt)'].sum().sort_values(ascending=False).head(20)

# Plot the data by report name and approved amount sort by total approved amount with title top 20 report names, include employee
top_20_report_names.plot(kind='bar', title='Top 20 Report Names by Total Approved Amount')

# Save the plot to Excel sheet
save_plot_to_sheet(wb, "Plot 3", top_20_report_names, "temp_plot3.png")

# print title and a blank line
#print("Plot 4 - Top 20 Expense Types by Total Approved Amount")
#print()

plot4 = 'Top 20 Expense Types by Total Approved Amount'

# Create the top 20 expense types dataframe
top_20_expense_types = filtered_df.groupby(['Expense Type'])['Approved Amount (rpt)'].sum().sort_values(ascending=False).head(20)

# Plot the data by expense type and approved amount sort by total approved amount with title top 20 expense types
top_20_expense_types.plot(kind='bar', title='Top 20 Expense Types by Total Approved Amount')

# Save the plot to Excel sheet
save_plot_to_sheet(wb, "Plot 4", top_20_expense_types, "temp_plot4.png")

# print title and a blank line
#print("Plot 5 - Top 20 Employees by Average Expense Amount")
#print()

plot5 = 'Top 20 Employees by Average Expense Amount'

# Create the top 20 employees by average approved amount dataframe
top_20_employees = filtered_df.groupby(['Employee'])['Approved Amount (rpt)'].mean().sort_values(ascending=False).head(20)

# Plot the average approved amount by employee id with title top 20 employees
top_20_employees.plot(kind='bar', title='Top 20 Employees by Average Approved Amount')

# Save the plot to Excel sheet
save_plot_to_sheet(wb, "Plot 5", top_20_employees, "temp_plot5.png")

# print title and a blank line
#print("Plot 6 - Top 20 Vendors by Average Expense Amount")
#print()

plot6 = 'Top 20 Vendors by Average Expense Amount'

# Create the top 20 vendors by average approved amount dataframe
top_20_vendors = filtered_df.groupby(['Vendor'])['Approved Amount (rpt)'].mean().sort_values(ascending=False).head(20)

# Plot the average approved amount by vendor with title top 20 vendors by average expense amount
top_20_vendors.plot(kind='bar', title='Top 20 Vendors by Average Expense Amount')

# Save the plot to Excel sheet
save_plot_to_sheet(wb, "Plot 6", top_20_vendors, "temp_plot6.png")

# print title and a blank line
#print("Plot 7 - Top 20 Report Names by Average Expense Amount")
#print()
plot7 = 'Top 20 Report Names by Average Expense Amount'

# Create the top 20 report names by average approved amount dataframe
top_20_report_names = filtered_df.groupby(['Report Name'])['Approved Amount (rpt)'].mean().sort_values(ascending=False).head(20)

# Plot the average approved amount by report name with title top 20 report names by average expense amount
top_20_report_names.plot(kind='bar', title='Top 20 Report Names by Average Expense Amount')

# Save the plot to Excel sheet
save_plot_to_sheet(wb, "Plot 7", top_20_report_names, "temp_plot7.png")


# print title and a blank line
#print("Plot 8 - Top 20 Expense Types by Average Expense Amount")
#print()

plot8 = 'Top 20 Expense Types by Average Expense Amount'

# Create the top 20 expense types by average approved amount dataframe
top_20_expense_types = filtered_df.groupby(['Expense Type'])['Approved Amount (rpt)'].mean().sort_values(ascending=False).head(20)

# Plot the average approved amount by expense type with title top 20 expense types by average expense amount
top_20_expense_types.plot(kind='bar', title='Top 20 Expense Types by Average Expense Amount')
plt.xticks(rotation=45, ha='right')
plt.xlabel('Expense Type')
plt.ylabel('Average Approved Amount (USD)')
plt.show()

# Save the plot to Excel sheet
save_plot_to_sheet(wb, "Plot 8", top_20_expense_types, "temp_plot8.png")

# Ensure 'Transaction Date' column is of datetime type
filtered_df['Transaction Date'] = pd.to_datetime(filtered_df['Transaction Date'])

# Group by month and year and calculate the total approved amount
total_approved_by_month = filtered_df.groupby([filtered_df['Transaction Date'].dt.year, filtered_df['Transaction Date'].dt.month])['Approved Amount (rpt)'].sum()

# Sort by year and month
total_approved_by_month = total_approved_by_month.sort_index()

# Rename the index to month names and years
total_approved_by_month.index = total_approved_by_month.index.set_names(['Year', 'Month'])
total_approved_by_month.index = total_approved_by_month.index.map(lambda x: (x[0], calendar.month_abbr[x[1]]))

# Create the plot for the total approved amount by month and year
total_approved_by_month.plot(kind='bar', title='Total Approved Amount by Month and Year')
plt.xticks(rotation=45, ha='right')
plt.xlabel('Month and Year')
plt.ylabel('Total Approved Amount (USD)')
plt.show()

# Create plot data
plot_data = total_approved_by_month

# Save the plot to Excel sheet
save_plot_to_sheet(wb, "Plot 9", plot_data, "temp_plot9.png")

# print title and a blank line
#print("Plot 9 - Total Approved Amount by Month")
#print()
# define plot 9
plot9 = 'Total Approved Amount by Month'

# print the total approved amount by month and year
#print(total_approved_by_month)

# Assuming you have grouped_df3 DataFrame already
# convert 'Transaction Date' column to datetime type
filtered_df['Transaction Date'] = pd.to_datetime(filtered_df['Transaction Date'])

# Group by month and year and calculate the average approved amount
avg_approved_by_month = filtered_df.groupby([filtered_df['Transaction Date'].dt.year, filtered_df['Transaction Date'].dt.month])['Approved Amount (rpt)'].mean()

# Sort by year and month
avg_approved_by_month = avg_approved_by_month.sort_index()

# Rename the index to month names and years
avg_approved_by_month.index = avg_approved_by_month.index.set_names(['Year', 'Month'])
avg_approved_by_month.index = avg_approved_by_month.index.map(lambda x: (x[0], calendar.month_abbr[x[1]]))
avg_approved_by_month.index = avg_approved_by_month.index.map(lambda x: f"{x[1]} {x[0]}")

# Create the plot for the average approved amount by month and year
avg_approved_by_month.plot(kind='bar', title='Average Approved Amount by Month and Year')
plt.xticks(rotation=45, ha='right')
plt.xlabel('Month and Year')
plt.ylabel('Average Approved Amount (USD)')
plt.show()

# Create plot data
plot_data = avg_approved_by_month

# Save the plot to Excel sheet
save_plot_to_sheet(wb, "Plot 10", plot_data, "temp_plot10.png")

# Define plot 10
plot10 = 'Average Approved Amount by Month'

# print title and a blank line
#print("Plot 10 - Average Approved Amount by Month")
#print()

# print the average approved amount by month and year
#print(avg_approved_by_month)

# plot the bar chart
avg_approved_by_month.plot(kind='bar', title='Average Approved Amount by Month')
plt.xlabel('Month')
plt.ylabel('Average Approved Amount (USD)')
plt.show()

# Convert 'Transaction Date' and 'Sent for Payment Date' columns to datetime type
filtered_df['Transaction Date'] = pd.to_datetime(filtered_df['Transaction Date'])
filtered_df['Sent for Payment Date'] = pd.to_datetime(filtered_df['Sent for Payment Date'])

# Calculate the time difference between transaction date and sent for payment date in seconds
filtered_df['Difference'] = (filtered_df['Sent for Payment Date'] - filtered_df['Transaction Date']).dt.total_seconds()

# Calculate average time difference between transaction date and sent for payment date by employee
avg_time_diff_by_employee = filtered_df.groupby(['Employee'])['Difference'].mean()

# Sort by average time difference in descending order and show only the top 20 employees
top_20 = avg_time_diff_by_employee.sort_values(ascending=False).head(20)

# Create a scatter plot with employee on the x-axis and average time difference on the y-axis, sorted by average time difference in descending order
plt.scatter(top_20.index, top_20.values)
plt.title('Average Time Difference between Transaction Date and Sent for Payment Date by Employee (Top 20)')
plt.xlabel('Employee')
plt.xticks(rotation=90)
plt.ylabel('Average Time Difference (Seconds)')
plt.show()

# Create plot data
plot_data = top_20

# Save the plot to Excel sheet
save_plot_to_sheet(wb, "Plot 11", plot_data, "temp_plot11.png")

# Define plot 11
plot11 = 'Average Time Difference between Transaction Date and Sent for Payment Date by Employee (Top 20)'

# Define plot 12
plot12 = 'Time Differences between Transaction Date and Sent for Payment Date'

# Create a DataFrame containing employee, transaction date, sent for payment date, and difference in days
df = filtered_df[['Employee', 'Transaction Date', 'Sent for Payment Date', 'Difference']]
df['Difference'] = df['Difference'] / (24*3600)  # Convert to days

# Create a scatter plot with employee on the x-axis and time difference in days on the y-axis
plt.scatter(df['Employee'], df['Difference'])
plt.title('Time Difference between Transaction Date and Sent for Payment Date')
plt.xlabel('Employee')
plt.xticks(rotation=90)
plt.ylabel('Time Difference (Days)')
plt.show()

# Save the plot to Excel sheet
save_plot_to_sheet(wb, "Plot 12", df, "temp_plot12.png")

# print a table of transactions with time differences over 180 days between transaction date and sent for payment date sort by difference in descending order, only show employee, transaction date, sent for payment date, and difference, show the difference in days
df = df[df['Difference'] > 180].sort_values(by='Difference', ascending=False)[['Employee', 'Transaction Date', 'Sent for Payment Date', 'Difference']]
df['Difference'] = df['Difference'] / (24*3600) # convert to days

# create plot data
plot_data = df

# save the table to Excel sheet
save_plot_to_sheet(wb, "Table 13", plot_data, "temp_table13.png")

# print title and a blank line
#print("Plot 13 - Time Differences over 180 Days between Transaction Date and Sent for Payment Date")
#print()
# define plot 13
plot13 = 'Time Differences over 180 Days between Transaction Date and Sent for Payment Date'



# print first 10    rows of the table   
#print(df.head(10))


 
#plot the time difference in days between transaction date and sent for payment date in a scatter plot only show the time differences greater than 180 days
plt.scatter(df['Employee'], df['Difference'])
plt.title('Time Difference between Transaction Date and Sent for Payment Date (Over 180 Days)')
plt.xlabel('Employee')
plt.xticks(rotation=90)
plt.ylabel('Time Difference (Days)')
plt.show()

result_df.plot(x='Employee ID', y='Approved Amount Difference', kind='bar', figsize=(10, 5), legend=None)
plt.title('Approved Amount Difference from Average of vendor by Employee ID')
plt.xlabel('Employee ID')
plt.ylabel('Approved Amount Difference')
plt.xticks(rotation=45, ha='right')
plt.tight_layout()
plt.savefig("temp_plot14.png")
plt.show()
# create plot data
plot_data = result_df

# save the plot to Excel sheet
save_plot_to_sheet(wb, "Plot 14", plot_data, "temp_plot14.png")

# print title and a blank line
#print("Plot 14 - Approved Amount Difference from Average of vendor by Employee ID")
#print()
# define plot 14
plot14 = 'Approved Amount Difference from Average of vendor by Employee ID'

# Display the result
#print(result_df)

# Scatter Plot by employee id, approved amount differences on y, employee id on x axis
plt.scatter(result_df['Employee ID'], result_df['Approved Amount Difference'])
plt.title('Approved Amount Difference by Employee ID')
plt.xlabel('Employee ID')
plt.ylabel('Approved Amount Difference (USD)')
plt.show()

# filter for difference is greater than 100
result_df = result_df[result_df['Approved Amount Difference'] > 100]

# Display the result
#print(result_df)

# create plot data
plot_data = result_df

# save the plot to Excel sheet
save_plot_to_sheet(wb, "Plot 15", plot_data, "temp_plot15.png")

# print title and a blank line
#print("Plot 15 - Approved Amount Difference from Average of Vendor (Over $100)")
#print()
# define plot 15
plot15 = 'Approved Amount Difference from Average of Vendor (Over $100)'

# Scatter Plot by employee id, approved amount differences on y, employee id on x axis, filter for difference is greater than 100
plt.scatter(result_df['Employee ID'], result_df['Approved Amount Difference'])
plt.title('Approved Amount Difference by Employee ID (Over $100)')
plt.xlabel('Employee ID')
plt.ylabel('Approved Amount Difference (USD)')
plt.show()

# print finished message
#print("Finished Saving Plots to Excel File")

# create a list of all the tests
all_tests = [test1, test2, test3, test4, test5, test6]

# create a list of all the plots
all_plots = [plot1, plot2, plot3, plot4, plot5, plot6, plot7, plot8, plot9, plot10, plot11, plot12, plot13, plot14, plot15]

# All descriptions = all_tests + all_plots
all_descriptions = all_tests + all_plots

# create cover sheet
create_cover_sheet(all_tests,all_plots,all_descriptions)

# save the workbook
wb.save(selected_file)

# close the workbook
wb.close()

# open the workbook
os.startfile(selected_file)