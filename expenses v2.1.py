import pandas as pd
import tkinter as tk
from tkinter import filedialog
import matplotlib.pyplot as plt
import calendar
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from fuzzywuzzy import fuzz
from fuzzywuzzy import process
# Custom formatting function
def format_currency(val):
    return "${:,.2f}".format(val)


def find_fuzzy_match(df, threshold=80):
    # Function to apply fuzzy matching on the Vendor column
    matched_df = pd.DataFrame()
    
    for employee_id, employee_group in df.groupby('Employee ID'):
        for vendor in employee_group['Vendor'].unique():
            vendor_group = employee_group[employee_group['Vendor'] == vendor]
            for _, row in vendor_group.iterrows():
                other_rows = vendor_group[vendor_group.index != row.name]
                for _, other_row in other_rows.iterrows():
                    if fuzz.token_set_ratio(row['Vendor'], other_row['Vendor']) >= threshold:
                        matched_df = matched_df.append(row)
                        matched_df = matched_df.append(other_row)
    
    return matched_df.drop_duplicates()

# create a cover sheet with links to all the tests and plots
def create_cover_sheet(tests, plots, descriptions):
    # Create a new workbook
    #wb = openpyxl.Workbook()

    # Create the cover sheet
    sheet = wb.active
    sheet.title = "Cover Sheet"

    # Add a title
    sheet['A1'] = "List of Tests and Plots"

    # Add the test names and plot links
    current_row = 3
    for i, test in enumerate(tests):
        # Add the test name as a hyperlink to the corresponding sheet, must be formatted blue and underlined
        sheet[f'A{current_row}'].value = f"Test {i+1}: {test}"
        sheet[f'A{current_row}'].hyperlink = f"'Test {i+1}'!A1"
        current_row += 1

    # Add the plot links
    for i, plot in enumerate(plots):
        # Add the plot name as a hyperlink to the corresponding sheet
        sheet[f'A{current_row}'].value = f"Plot {i+1}: {plot}"
        sheet[f'A{current_row}'].hyperlink = f"'Plot {i+1}'!A1"
        current_row += 1

    # Resize columns to fit data
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    # Add descriptions to all tabs and add links back to the cover sheet
    for ws in wb.worksheets:
        if ws.title.startswith("Test "):
            test_num = int(ws.title.split()[1])
            ws['A1'] = f"Test {test_num}: {tests[test_num - 1]}"
        elif ws.title.startswith("Plot "):
            plot_num = int(ws.title.split()[1])
            ws['A1'] = f"Plot {plot_num}: {plots[plot_num - 1]}"

            # Add a link back to the cover sheet
            
        ws['A1'].hyperlink = "'Cover Sheet'!A1"

    return wb

# create a function to define fuzzy matches
def within_1_percent(df):
    """
    Filter data within 1% of each other in the Approved Amount 
    """
    if len(df) >= 2:
        mean_payment = df['Approved Amount (rpt)'].mean()
        upper_limit = mean_payment * 1.01
        lower_limit = mean_payment * 0.99
        return df[(df['Approved Amount (rpt)'] <= upper_limit) & (df['Approved Amount (rpt)'] >= lower_limit)]
    else:
        return pd.DataFrame()
def within_5_days(df):
    """
    Filter data for transactions that happened within 5 days of each other
    """
    if len(df) >= 2:
        max_date = df['Transaction Date'].max()
        min_date = df['Transaction Date'].min()
        if (max_date - min_date).days <= 5:
            return df
        else:
            return pd.DataFrame()
    else:
        return pd.DataFrame()

# Define a function to filter groups with less than 2 rows
def at_least_2_rows(df):
    return len(df) >= 2
def select_file():
    root = tk.Tk()
    root.withdraw()

    while True:
        print("Select a file to load for analysis")
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xls;*.xlsx;*.xlsm;*.xlsb;*.odf;*.ods;*.odt")])

        if file_path:
            print("File selected: " + file_path)
            break
        else:
            print("No file selected. Please try again.")

    return file_path

file_path = select_file()
df = pd.read_excel(file_path, skiprows=2, engine='openpyxl')

def select_save_file():
    """
    Opens a Windows dialog for the user to select a file name and location
    to save the results as an Excel file.

    Returns:
        str: The selected file path, or None if no file was selected.
    """
    print("Select a file to save the results")
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        title="Select a file to save the results",
        filetypes=[("Excel Files", "*.xlsx")]
    )
    return file_path if file_path else None

# Function to save DataFrame to Excel sheet
def save_df_to_sheet(workbook, sheet_name, df):
    try:
        ws = workbook.create_sheet(sheet_name)
        for r in dataframe_to_rows(df, index=True, header=True):
            ws.append(r)
        
        # Resize columns to fit data
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    except Exception as e:
        print(f"An error occurred while saving DataFrame to sheet '{sheet_name}': {e}")



# Function to save plot to Excel sheet
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image

from openpyxl.utils.dataframe import dataframe_to_rows

def save_plot_to_sheet(workbook, sheet_name, plot_data, table_data, img_name):
    try:
        fig, ax = plt.subplots()
        # plot data, title = test or plot description
        plot_data.plot(kind='bar', title='', ax=ax)
        fig.savefig(img_name, dpi=100)
        
        ws = workbook.create_sheet(sheet_name)
        img = Image(img_name)
        ws.add_image(img, 'A2')
        
        # save the table_data DataFrame starting in cell L1
        rows = dataframe_to_rows(table_data.reset_index(), index=False, header=True)

        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 18):  # Change 1 to 12 to start from the L column
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value
        
        # Resize columns to fit data
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    except Exception as e:
        print(f"An error occurred while saving plot to sheet '{sheet_name}': {e}")


        
    

# Example usage:
selected_file = select_save_file()
# Create a new workbook
wb = Workbook()
if selected_file:
    print(f"Selected file: {selected_file}")
else:
    print("No file selected.")

# Convert the columns to formats we need
df['Transaction Date'] = pd.to_datetime(df['Transaction Date'])
df['Approved Amount (rpt)'] = pd.to_numeric(df['Approved Amount (rpt)'])
df['Sent for Payment Date'] = pd.to_datetime(df['Sent for Payment Date'])


# approved amount is a string, so we need to convert it to a currency with 2 decimal places and commas
df['Approved Amount (rpt)'] = pd.to_numeric(df['Approved Amount (rpt)'])



df['Sent for Payment Date'] = pd.to_datetime(df['Sent for Payment Date'])

# Filter data for transactions over 1000 and sort by employee ID  
filtered_df = df[df['Approved Amount (rpt)'] > 25].sort_values(by=['Employee ID'])

#sort by employee ID
filtered_df = filtered_df.sort_values(by=['Employee ID'])

#limit to 100. rows
#filtered_df = filtered_df.head(1000)

# TEST 1-------------------Same employee, same vendor same total
#define test1
test1 = 'Same employee, same vendor same total'

# Group by employee id, vendor, and approved amount
grouped_df = filtered_df.groupby(['Employee ID', 'Vendor', 'Approved Amount (rpt)'])

# Filter groups with more than one row per employee
filtered_groups = grouped_df.filter(lambda x: len(x) > 1)

# Sort the filtered groups by the approved amount in descending order, then by employee id, vendor, and transaction date
sorted_filtered_groups = filtered_groups.sort_values(by=['Approved Amount (rpt)', 'Employee ID', 'Vendor', 'Transaction Date'], ascending=[False, True, True, True])

# Save the grouped data to a new sheet in the Excel file
save_df_to_sheet(wb, "Test 1", sorted_filtered_groups[['Employee','Employee ID','Transaction Date','Approved Amount (rpt)','Transaction Date','Vendor','Report Name']])
# TEST 2-----------same employee, amount and date -----------------------------
test2 = 'Same employee, amount and date'

# Group by employee id, approved amount, and transaction date
grouped_df2 = filtered_df.groupby(['Employee ID', 'Approved Amount (rpt)', 'Transaction Date'])

# Filter groups with more than one row per employee
filtered_groups2 = grouped_df2.filter(lambda x: len(x) > 1)

# Sort the filtered groups by approved amount (descending), employee id, and transaction date
sorted_filtered_groups2 = filtered_groups2.sort_values(by=['Approved Amount (rpt)', 'Employee ID', 'Transaction Date'], ascending=[False, True, True])

# Save the grouped data to a new sheet in the Excel file
save_df_to_sheet(wb, "Test 2", sorted_filtered_groups2[['Employee','Employee ID','Transaction Date','Approved Amount (rpt)','Transaction Date','Vendor','Report Name']])
test3 = 'Same employee, same amount'

# Group by employee id and approved amount
grouped_df3 = filtered_df.groupby(['Employee ID', 'Approved Amount (rpt)'])

# Filter groups with more than one row per employee
filtered_groups3 = grouped_df3.filter(lambda x: len(x) > 1)

# Sort the filtered groups by approved amount (descending) and employee id
sorted_filtered_groups3 = filtered_groups3.sort_values(by=['Approved Amount (rpt)', 'Employee ID'], ascending=[False, True])

# Save the grouped data to a new sheet in the Excel file
save_df_to_sheet(wb, "Test 3", sorted_filtered_groups3[['Employee','Employee ID','Transaction Date','Approved Amount (rpt)','Transaction Date','Vendor','Report Name']])


# Custom function to find the matched keyword in the report name
def find_matched_keyword(report_name, keywords):
    for keyword in keywords:
        if keyword in report_name:
            return keyword
    return None

test4 = 'Keyword test'

# Create a Tkinter file dialog box to open an Excel file
root = tk.Tk()
root.withdraw()
print("Please select keywords file")
file_path = filedialog.askopenfilename()
print("Loading Keywords: ", file_path, "")
import re

# Read in keywords from the Excel file
keywords = pd.read_excel(file_path)

# Convert the first column of keywords DataFrame to a list
keywords = keywords.iloc[:, 0].tolist()

def find_matched_keyword(text, keywords):
    for keyword in keywords:
        if re.search(r'\b' + re.escape(keyword) + r'\b', text, re.IGNORECASE):
            return keyword
    return None

# Create a regular expression pattern with word boundaries
pattern = '|'.join(r'\b' + re.escape(keyword) + r'\b' for keyword in keywords)

# Group the data with any of the keywords, sort by approved amount (descending) and keyword
grouped_df4 = filtered_df[filtered_df['Report Name'].str.contains(pattern, regex=True, na=False)].sort_values(by=['Approved Amount (rpt)', 'Report Name'], ascending=[False, True])

# Remove empty dataframes from the grouped data
grouped_df4 = grouped_df4[grouped_df4['Approved Amount (rpt)'].notna()]

# Add a column to the grouped_df4 DataFrame indicating the matched keyword
grouped_df4['Matched Keyword'] = grouped_df4['Report Name'].apply(lambda x: find_matched_keyword(x, keywords))

# Save the grouped data to a new sheet in the Excel file, including the 'Matched Keyword' column
save_df_to_sheet(wb, "Test 4", grouped_df4[['Employee', 'Employee ID', 'Transaction Date', 'Report Name', 'Approved Amount (rpt)', 'Matched Keyword']])


# Define test5
test5 = 'Same employee and fuzzy match on vendor, date and amount'

grouped_df5 = filtered_df.groupby(['Employee ID', 'Vendor']).apply(lambda x: within_1_percent(within_5_days(x)))

# Reset the index after applying the filter functions
grouped_df5.reset_index(drop=True, inplace=True)

# Filter out groups with less than 2 rows
grouped_df5 = grouped_df5[grouped_df5.groupby(['Employee ID', 'Vendor']).transform(at_least_2_rows)['Approved Amount (rpt)']]

# Reset the index again
grouped_df5.reset_index(drop=True, inplace=True)

# Apply fuzzy matching on vendors
fuzzy_matched_df5 = find_fuzzy_match(grouped_df5)

# Save the grouped data to a new sheet in the Excel file
save_df_to_sheet(wb, "Test 5", fuzzy_matched_df5[['Employee','Employee ID','Transaction Date','Approved Amount (rpt)','Transaction Date','Vendor','Report Name']])

test6 = 'Same employee and expense type fuzzy match on date and amount'

grouped_df6 = filtered_df.groupby(['Employee ID', 'Expense Type']).apply(lambda x: within_1_percent(within_5_days(x)))

# Reset the index after applying the filter functions
grouped_df6.reset_index(drop=True, inplace=True)

# Filter out groups with less than 2 rows
grouped_df6 = grouped_df6[grouped_df6.groupby(['Employee ID', 'Expense Type']).transform(at_least_2_rows)['Approved Amount (rpt)']]

# Reset the index again
grouped_df6.reset_index(drop=True, inplace=True)

# Sort by tuples with the highest approved amount
grouped_df6 = grouped_df6.sort_values(by=['Approved Amount (rpt)'], ascending=False)

# Save the grouped data to a new sheet in the Excel file
save_df_to_sheet(wb, "Test 6", grouped_df6[['Employee','Employee ID','Transaction Date','Approved Amount (rpt)','Transaction Date','Vendor','Report Name']])

# TEST 7-------------------Different employees, same amount, same date
test7 = 'Different employees, same amount, same date'

# Group by transaction date, approved amount, and vendor
grouped_df7 = filtered_df.groupby(['Transaction Date', 'Approved Amount (rpt)', 'Vendor'])

# Filter groups with more than one unique employee
filtered_groups7 = grouped_df7.filter(lambda x: x['Employee ID'].nunique() > 1)

# Sort the filtered groups by approved amount (descending), transaction date, and vendor
sorted_filtered_groups7 = filtered_groups7.sort_values(by=['Approved Amount (rpt)', 'Transaction Date', 'Vendor'], ascending=[False, True, True])

# Save the grouped data to a new sheet in the Excel file
save_df_to_sheet(wb, "Test 7", sorted_filtered_groups7[['Employee','Employee ID','Transaction Date','Approved Amount (rpt)','Transaction Date','Vendor','Report Name']])

# define plot 1
plot1 = 'Top 20 Employees by Total Approved Amount'


# print the top 20 employees by total approved amount, include number of transactions and average amount
top_20_employees = filtered_df.groupby(['Employee'])['Approved Amount (rpt)'].agg(['sum', 'count', 'mean']).sort_values(by='sum', ascending=False).head(20)

# Create plot data, excluding count and mean
plot_data = top_20_employees.drop(['count', 'mean'], axis=1)

#sum and mean should be rounded to 2 decimal places and have commas and dollar signs
top_20_employees['sum'] = top_20_employees['sum'].map('${:,.2f}'.format)
top_20_employees['mean'] = top_20_employees['mean'].map('${:,.2f}'.format)

table_data = top_20_employees
# Save the plot to Excel sheet
save_plot_to_sheet(wb, "Plot 1", plot_data, table_data, "temp_plot1.png")

# define plot 2
plot2 = 'Top 20 Vendors by Total Approved Amount'

# Create the top 20 vendors dataframe, including number of transactions and average amount
top_20_vendors = filtered_df.groupby(['Vendor'])['Approved Amount (rpt)'].agg(['sum', 'count', 'mean']).sort_values(by='sum', ascending=False).head(20)

#drop count and mean
plot_data = top_20_vendors.drop(['count', 'mean'], axis=1)

#sum and mean should be rounded to 2 decimal places and have commas and dollar signs
top_20_vendors['sum'] = top_20_vendors['sum'].map('${:,.2f}'.format)
top_20_vendors['mean'] = top_20_vendors['mean'].map('${:,.2f}'.format)

table_data = top_20_vendors

# Save the plot to Excel sheet
save_plot_to_sheet(wb, "Plot 2", plot_data, table_data, "temp_plot2.png")

# define plot 3
plot3 = 'Top 20 Report Names by Total Approved Amount'

# Create the top 20 report names dataframe including number of transactions and average amount
top_20_report_names = filtered_df.groupby(['Report Name'])['Approved Amount (rpt)'].agg(['sum', 'count', 'mean']).sort_values(by='sum', ascending=False).head(20)

#drop count and mean
plot_data = top_20_report_names.drop(['count', 'mean'], axis=1)

#sum and mean should be rounded to 2 decimal places and have commas and dollar signs
top_20_report_names['sum'] = top_20_report_names['sum'].map('${:,.2f}'.format)
top_20_report_names['mean'] = top_20_report_names['mean'].map('${:,.2f}'.format)

table_data = top_20_report_names

# Save the plot to Excel sheet
save_plot_to_sheet(wb, "Plot 3", plot_data, table_data, "temp_plot3.png")

plot4 = 'Top 20 Expense Types by Total Approved Amount'

# Create the top 20 expense types dataframe, including number of transactions and average amount
top_20_expense_types = filtered_df.groupby(['Expense Type'])['Approved Amount (rpt)'].agg(['sum', 'count', 'mean']).sort_values(by='sum', ascending=False).head(20)

# drop count and mean
plot_data = top_20_expense_types.drop(['count', 'mean'], axis=1)

#sum and mean should be rounded to 2 decimal places and have commas and dollar signs
top_20_expense_types['sum'] = top_20_expense_types['sum'].map('${:,.2f}'.format)
top_20_expense_types['mean'] = top_20_expense_types['mean'].map('${:,.2f}'.format)

table_data = top_20_expense_types

# Save the plot to Excel sheet
save_plot_to_sheet(wb, "Plot 4", plot_data, table_data, "temp_plot4.png")

plot5 = 'Top 20 Employees by Average Expense Amount'

# Create the top 20 employees by average approved amount dataframe including number of transactions and average amount
top_20_employees = filtered_df.groupby(['Employee'])['Approved Amount (rpt)'].agg(['sum', 'count', 'mean']).sort_values(by='mean', ascending=False).head(20)

# drop sum, count 
plot_data = top_20_employees.drop(['sum', 'count'], axis=1)

#sum and mean should be rounded to 2 decimal places and have commas and dollar signs
top_20_employees['sum'] = top_20_employees['sum'].map('${:,.2f}'.format)
top_20_employees['mean'] = top_20_employees['mean'].map('${:,.2f}'.format)

table_data = top_20_employees

# Plot the average approved amount by employee id with title top 20 employees
#plot_data.plot(kind='bar', title='Top 20 Employees by Average Expense Amount')

# Save the plot to Excel sheet
save_plot_to_sheet(wb, "Plot 5", plot_data, table_data, "temp_plot5.png")

# print title and a blank line
#print("Plot 6 - Top 20 Vendors by Average Expense Amount")
#print()

plot6 = 'Top 20 Vendors by Average Expense Amount'

# Create the top 20 vendors by average approved amount dataframe including number of transactions and average amount
top_20_vendors = filtered_df.groupby(['Vendor'])['Approved Amount (rpt)'].agg(['sum', 'count', 'mean']).sort_values(by='mean', ascending=False).head(20)

# drop sum, count 
plot_data = top_20_vendors.drop(['sum', 'count'], axis=1)

#sum and mean should be rounded to 2 decimal places and have commas and dollar signs
top_20_vendors['sum'] = top_20_vendors['sum'].map('${:,.2f}'.format)
top_20_vendors['mean'] = top_20_vendors['mean'].map('${:,.2f}'.format)

table_data = top_20_vendors

# Save the plot to Excel sheet
save_plot_to_sheet(wb, "Plot 6", plot_data, table_data, "temp_plot6.png")

plot7 = 'Top 20 Report Names by Average Expense Amount'

# Create the top 20 report names by average approved amount dataframe including number of transactions and average amount
top_20_report_names = filtered_df.groupby(['Report Name'])['Approved Amount (rpt)'].agg(['sum', 'count', 'mean']).sort_values(by='mean', ascending=False).head(20)

# drop sum, count
plot_data = top_20_report_names.drop(['sum', 'count'], axis=1)

#sum and mean should be rounded to 2 decimal places and have commas and dollar signs
top_20_report_names['sum'] = top_20_report_names['sum'].map('${:,.2f}'.format)
top_20_report_names['mean'] = top_20_report_names['mean'].map('${:,.2f}'.format)

table_data = top_20_report_names

# Save the plot to Excel sheet
save_plot_to_sheet(wb, "Plot 7", plot_data, table_data, "temp_plot7.png")

plot8 = 'Top 20 Expense Types by Average Expense Amount'

# Create the top 20 expense types by average approved amount dataframe, including number of transactions and total amount
top_20_expense_types = filtered_df.groupby(['Expense Type'])['Approved Amount (rpt)'].agg(['sum', 'count', 'mean']).sort_values(by='mean', ascending=False).head(20)

# drop count and sum
plot_data = top_20_expense_types.drop(['count', 'sum'], axis=1)

#sum and mean should be rounded to 2 decimal places and have commas and dollar signs
top_20_expense_types['sum'] = top_20_expense_types['sum'].map('${:,.2f}'.format)
top_20_expense_types['mean'] = top_20_expense_types['mean'].map('${:,.2f}'.format)


table_data = top_20_expense_types

# Save the plot to Excel sheet
save_plot_to_sheet(wb, "Plot 8", plot_data, table_data, "temp_plot8.png")

# Ensure 'Transaction Date' column is of datetime type
filtered_df['Transaction Date'] = pd.to_datetime(filtered_df['Transaction Date'])

# Group by month and year and calculate the total approved amount, include number of transactions and average amount
total_approved_by_month = filtered_df.groupby([filtered_df['Transaction Date'].dt.year, filtered_df['Transaction Date'].dt.month])['Approved Amount (rpt)'].agg(['sum', 'count', 'mean'])


# Sort by year and month
total_approved_by_month = total_approved_by_month.sort_index()

# Rename the index to month names and years
total_approved_by_month.index = total_approved_by_month.index.set_names(['Year', 'Month'])
total_approved_by_month.index = total_approved_by_month.index.map(lambda x: (x[0], calendar.month_abbr[x[1]]))

# Drop count and mean
plot_data = total_approved_by_month.drop(['count', 'mean'], axis=1)

#sum and mean should be rounded to 2 decimal places and have commas and dollar signs
total_approved_by_month['sum'] = total_approved_by_month['sum'].map('${:,.2f}'.format)
total_approved_by_month['mean'] = total_approved_by_month['mean'].map('${:,.2f}'.format)

table_data = total_approved_by_month

# Save the plot to Excel sheet
save_plot_to_sheet(wb, "Plot 9", plot_data, table_data, "temp_plot9.png")

# define plot 9
plot9 = 'Total Approved Amount by Month'

# convert 'Transaction Date' column to datetime type
filtered_df['Transaction Date'] = pd.to_datetime(filtered_df['Transaction Date'])

# Group by month and year and calculate the average approved amount, include number of transactions and total amount
avg_approved_by_month = filtered_df.groupby([filtered_df['Transaction Date'].dt.year, filtered_df['Transaction Date'].dt.month])['Approved Amount (rpt)'].agg(['sum', 'count', 'mean'])

# Sort by year and month
avg_approved_by_month = avg_approved_by_month.sort_index()

# Rename the index to month names and years
avg_approved_by_month.index = avg_approved_by_month.index.set_names(['Year', 'Month'])
avg_approved_by_month.index = avg_approved_by_month.index.map(lambda x: (x[0], calendar.month_abbr[x[1]]))
avg_approved_by_month.index = avg_approved_by_month.index.map(lambda x: f"{x[1]} {x[0]}")

# Drop count and sum
plot_data = avg_approved_by_month.drop(['count', 'sum'], axis=1)

#sum and mean should be rounded to 2 decimal places and have commas and dollar signs
avg_approved_by_month['sum'] = avg_approved_by_month['sum'].map('${:,.2f}'.format)
avg_approved_by_month['mean'] = avg_approved_by_month['mean'].map('${:,.2f}'.format)

table_data = avg_approved_by_month


# Save the plot to Excel sheet
save_plot_to_sheet(wb, "Plot 10", plot_data, table_data, "temp_plot10.png")

# Define plot 10
plot10 = 'Average Approved Amount by Month'

# Convert 'Transaction Date' and 'Sent for Payment Date' columns to datetime type
filtered_df['Transaction Date'] = pd.to_datetime(filtered_df['Transaction Date'])
filtered_df['Sent for Payment Date'] = pd.to_datetime(filtered_df['Sent for Payment Date'])

# Calculate the time difference between transaction date and sent for payment date in days
filtered_df['Difference'] = (filtered_df['Sent for Payment Date'] - filtered_df['Transaction Date']).dt.days

# Calculate average time difference between transaction date and sent for payment date by employee
avg_time_diff_by_employee = filtered_df.groupby(['Employee'])['Difference'].mean()

# Sort by average time difference in descending order and show only the top 20 employees
top_20 = avg_time_diff_by_employee.sort_values(ascending=False).head(20)

# Convert the index into a DataFrame
top_20_df = top_20.reset_index()

# Create a DataFrame with 'Employee' and 'Difference' columns
plot_data = top_20_df[['Employee', 'Difference']]

# format the data for the table to integers
top_20_df['Difference'] = top_20_df['Difference'].map('{:,.0f}'.format)

table_data = top_20_df

# Save the plot to Excel sheet
save_plot_to_sheet(wb, "Plot 11", plot_data, table_data, "temp_plot11.png")

# Define plot 11
plot11 = 'Average Time Difference between Transaction Date and Sent for Payment Date by Employee (Top 20)'

# Create a DataFrame containing employee, transaction date, sent for payment date, and difference in days and approved amount
df = filtered_df[['Employee', 'Transaction Date', 'Sent for Payment Date', 'Difference', 'Approved Amount (rpt)']]
#df['Difference'] = df['Difference'] / (24*3600)  # Convert to days

# print a table of transactions with time differences over 180 days between transaction date and sent for payment date sort by difference in descending order, only show employee, transaction date, sent for payment date, and difference, show the difference in days
df = df[df['Difference'] > 180].sort_values(by='Difference', ascending=False)[['Employee', 'Transaction Date', 'Sent for Payment Date', 'Difference', 'Approved Amount (rpt)']]

# create a dataframe with employee and difference in days
plot_data = df[['Employee', 'Difference']]

#convert approved amount to 2 decimal places and have commas and dollar signs
df['Approved Amount (rpt)'] = df['Approved Amount (rpt)'].map('${:,.2f}'.format)


# format the difference column to 2 decimal places and have commas and dollar signs
df['Difference'] = df['Difference'].map('{:,.0f}'.format)

# create a dataframe with employee, transaction date, sent for payment date, and difference in days
table_data = df[['Employee', 'Transaction Date', 'Sent for Payment Date', 'Difference', 'Approved Amount (rpt)']]

# save the plot to Excel sheet
save_plot_to_sheet(wb, "Plot 12", plot_data, table_data, "temp_plot13.png")

# define plot 12
plot12 = 'Time Differences over 180 Days between Transaction Date and Sent for Payment Date'

# Calculate the average approved amount per vendor
avg_approved_by_vendor = filtered_df.groupby(['Vendor'])['Approved Amount (rpt)'].mean()

# Add the vendor's average approved amount to the existing DataFrame
filtered_df['Vendor Average'] = filtered_df['Vendor'].map(avg_approved_by_vendor)

# Calculate the difference for each transaction by comparing the approved amount with the vendor's average
filtered_df['Difference'] = filtered_df['Approved Amount (rpt)'] - filtered_df['Vendor Average']

# Filter for differences greater than 100 and sort by difference in descending order
filtered_df = filtered_df[filtered_df['Difference'] > 100].sort_values(by='Difference', ascending=False)

# Create plot data with 'Employee' and 'Difference'
plot_data = filtered_df[['Employee', 'Difference']]


# Create table data with 'Employee', 'Vendor', 'Vendor Average', 'Approved Amount (rpt)', and 'Difference'
table_data = filtered_df[['Employee', 'Vendor', 'Vendor Average', 'Approved Amount (rpt)', 'Difference']]

#format the data for the table
table_data['Vendor Average'] = table_data['Vendor Average'].map('{:,.2f}'.format)
table_data['Approved Amount (rpt)'] = table_data['Approved Amount (rpt)'].map('{:,.2f}'.format)
table_data['Difference'] = table_data['Difference'].map('{:,.0f}'.format)

# define plot 13
plot13 = 'Approved Amount Difference from Average of Vendor (Over $100)'

# save the plot to Excel sheet
save_plot_to_sheet(wb, "Plot 13", plot_data, table_data, "temp_plot13.png")

# create a list of all the tests
all_tests = [test1, test2, test3, test4, test5, test6, test7]

# create a list of all the plots
all_plots = [plot1, plot2, plot3, plot4, plot5, plot6, plot7, plot8, plot9, plot10, plot11, plot12, plot13]

# All descriptions = all_tests + all_plots
all_descriptions = all_tests + all_plots

# create cover sheet
create_cover_sheet(all_tests,all_plots,all_descriptions)

# save the workbook
wb.save(selected_file)

# close the workbook
wb.close()

# open the workbook
os.startfile(selected_file)